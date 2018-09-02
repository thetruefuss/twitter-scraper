import csv  # To read & write Comma Seperated Value files
import json  # To encode python objects to json
import os
import re  # For regular expressions
import sys
import time  # To manipulate time values
import warnings
from datetime import datetime  # To parse datetime types
from multiprocessing import Process, Queue  # Analogous to threading but uses processes & provides simple API for dummy subpackage
from multiprocessing.dummy import Lock
from multiprocessing.dummy import Pool as ThreadPool
from multiprocessing.dummy import Queue as ThreadQueue
from pprint import pprint  # To pretty print different data structures in python

import requests
from dateutil import parser as dateparser
from openpyxl import load_workbook

import settings
import sqlalchemy.exc
from pyquery import PyQuery
from sqlalchemy import *
from sqlalchemy import event, exc
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import relationship, scoped_session, sessionmaker

time_wait = 0
flag1 = False

"""
REGEX => https://github.com/praritlamba/Mining-Twitter-Data-for-Sentiment-Analysis/blob/master/README.md
"""

emoticons_str = r"""
    (?:
        [:=;] # Eyes
        [oO\-]? # Nose (optional)
        [D\)\]\(\]/\\OpP] # Mouth
    )"""

regex_str = [
    emoticons_str,
    r'<[^>]+>',  # HTML tags
    r'(?:@[\w_]+)',  # @-mentions
    r"(?:\#+[\w_]+[\w\'_\-]*[\w_]+)",  # hash-tags
    r"(?:\$+[a-zA-Z]+[\w\'_\-]*[\w_]+)",  # cash-tags
    r'http[s]?://(?:[a-z]|[0-9]|[$-_@.&amp;+]|[!*\(\),]|(?:%[0-9a-f][0-9a-f]))+',  # URLs

    r'(?:(?:\d+,?)+(?:\.?\d+)?)',  # numbers
    r"(?:[a-z][a-z'\-_]+[a-z])",  # words with - and '
    r'(?:[\w_]+)',  # other words
    r'(?:\S)'  # anything else
]

tokens_re = re.compile(r'(' + '|'.join(regex_str) + ')', re.VERBOSE | re.IGNORECASE)
emoticon_re = re.compile(r'^' + emoticons_str + '$', re.VERBOSE | re.IGNORECASE)


class LoadingError(Exception):
    pass

# Get tweets
class Twit:
    def __init__(self):
        pass

    def json(self):
        return {'date': self.unixtime, 'text': self.text, 'screen_name': self.screen_name, 'user_name': self.user_name,
                'user_id': self.user_id,
                'id': self.id, 'retweets_count': self.retweets_count, 'favorites_count': self.favorites_count,
                'permalink': self.permalink, 'urls': self.urls, 'mentions': self.mentions,
                'hashtags': self.hashtags, 'is_retweet': self.is_retweet, 'symbols': self.symbols,
                'is_protected': self.is_protected}

    def __repr__(self):
        return self.text


class Page(object):
    """Starts a requests session with twitter."""

    def __init__(self, proxy=None):
        self.pr = { 'http': proxy, 'https': proxy } # for sending request using proxy, if twitter blocks your current ip
        self.timeout = 30 # timeout request if not getting response till 30 seconds
        self.ses = requests.Session() # starting a session, so we could persist certain parameters across requests
        # using dict for sending some headers in our request
        self.ses.headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.111 Safari/537.36',
            'Connection': 'keep-alive',
            'Cache-Control': 'max-age=0',
            'Accept-Encoding': 'gzip, deflate, sdch, br',
            # 'accept': 'application/json, text/javascript, */*; q=0.01',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US;q=0.6,en;q=0.4',
            'x-compress': 'null',
            'Upgrade-Insecure-Requests': '1',
            # 'x-requested-with': 'XMLHttpRequest',
            # 'x-twitter-active-user': 'yes',
            'host': 'twitter.com'
        }

    def __del__(self):
        self.ses.close()

    def close(self):
        """
        Close the session.

        :return:
        """
        self.ses.close()

    def load(self, url, params=None, headers=None, important=True):
        """Returns request content after getting response from given url & params(GET), headers and timeout args."""
        error_count = 0 # counting request failures

        while True:
            try:
                resp = self.ses.get(url, proxies=self.pr, params=params, headers=headers, timeout=self.timeout)
            except requests.exceptions.RequestException as e:
                error_count += 1
                # print('Loading error', error_count, pr, e)
                if error_count < 3:
                    print('Loading error', error_count, url, self.pr, e)
                    time.sleep(60)
                    continue
                print('Error limit exceeded Loading error', url, self.pr, e)
                if important:
                    raise LoadingError
                else:
                    return None

            # showing different error messages depending upon different status_code(s), if request was important enough then raise LoadingError(Exception)
            if resp.status_code == requests.codes.ok:
                return resp.text

            elif resp.status_code == 404:
                print('Error 404', url, resp.text, resp.status_code)
                if important:
                    raise LoadingError
                else:
                    return None

            elif resp.status_code == 429:
                print('Rate limit. Sleep 3 min')
                time.sleep(3 * 60)
                continue

            elif resp.status_code == 503:
                # print('503 Error waiting 2 min', screen_name, proxy, url, resp.text, resp.status_code)
                error_count += 1
                if error_count > 5:
                    print('AWAM: Requestes 503 error', url, self.pr)
                    if important:
                        raise LoadingError
                    else:
                        return None
                print(' Requestes 503 error', error_count, url, self.pr)
                time.sleep(120)
                continue

            else:
                print('Error', url, resp.text, resp.status_code)
                error_count += 1
                print('Loading error', error_count)
                if error_count > 5:
                    print('Error limit exceeded Requestes error ', url, self.pr, resp.status_code)
                    if important:
                        raise LoadingError
                    else:
                        return None
                print('Error waiting 1 min', url, resp.text, resp.status_code)
                time.sleep(60)
                continue




def twitter_login(page, login, password):
    """Function for twitter login."""
    resp = page.load('https://twitter.com/')

    res = re.findall('<input type="hidden" value="(.+?)" name="authenticity_token">', resp) # getting authenticity_token, just like csrf_token in django for verifying request
    token = res[0]

    params = {'session[username_or_email]': login,
              'session[password]': password,
              'remember_me': '1',
              'return_to_ssl': 'true',
              'scribe_log': '',
              'redirect_after_login': '/',
              'authenticity_token': token}

    url = 'https://twitter.com/sessions'
    while True:
        resp = page.ses.post(url, data=params, timeout=10)

        if resp.status_code == requests.codes.ok:
            if re.search('action="/logout" method="POST">', resp.text):
                print('Logged as', login)
                res = re.findall('<input type="hidden" value="(.+?)" name="authenticity_token', resp.text)
                token = res[0]
                return token

            elif re.search('Your account appears to have exhibited automated behavior that violates', resp.text):
                print('Your account appears to have exhibited automated behavior that violates')
                print('Pass a Google reCAPTCHA challenge.Verify your phone number')
                return False

            elif re.search('id="login-challenge-form"', resp.text):
                authenticity_token = re.findall('name="authenticity_token" value="(.+?)"/>', resp.text, re.S)[0]
                challenge_id = re.findall('name="challenge_id" value="(.+?)"/>', resp.text, re.S)[0]
                user_id = re.findall('name="user_id" value="(.+?)"/>', resp.text, re.S)[0]
                challenge_type = re.findall('name="challenge_type" value="(.+?)"/>', resp.text, re.S)[0]
                platform = re.findall('name="platform" value="(.+?)"/>', resp.text, re.S)[0]

                params = {
                    'authenticity_token': authenticity_token,
                    'challenge_id': challenge_id,
                    'user_id': user_id,
                    'challenge_type': challenge_type,
                    'platform': platform,
                    'redirect_after_login': '/',
                    'remember_me': 'true',
                    'challenge_response': '+1-812-679-7792'}
                url = 'https://twitter.com/account/login_challenge'
                print(challenge_type)
                continue

            elif re.search('You have initiated too many login verification requests', resp.text, re.S):
                print('You have initiated too many login verification requests')
                global time_wait
                print(time_wait)
                time_wait = time.time()
                raise LoadingError

            else:
                print('Not logged as', login)
                return False
            break

        else:
            print('Login Error', resp.status_code)
            return False
    # resp = page.ses.post('https://twitter.com/sessions', data=params, timeout=10)
    #
    # if re.search('action="/logout" method="POST">', resp.text):
    #     print('Logged as', login)
    #
    #     res = re.findall('<input type="hidden" value="(.+?)" name="authenticity_token', resp.text)
    #     token = res[0]
    #     return token
    # else:
    #     print('Not logged as', login)
    #     return None




def get_new_search(page, query, login=None, password=None, nativeretweets=False):
    """Searches for given query and returns a list of all tweets found."""
    user_name = query[0]
    query_string = query[1]
    data_begin = query[2]
    data_end = query[3]
    refreshCursor = ''

    h = {'x-requested-with': 'XMLHttpRequest', 'x-twitter-active-user': 'yes',
         'accept': 'application/json, text/javascript, */*; q=0.01'}

    params = {}
    # if we want to scrape our own tweets, first login then start scraping
    if settings.PROFILE_SEARCH:
        refreshCursor = '999992735314882560'
        token = twitter_login(page, settings.TWITTER_USERNAME, settings.TWITTER_PASSWORD)
        if not token:
            print('UNABLE LOGIN')
            return False

    empty_count = 0
    date_range_change_count = 0
    while True:
        if settings.PROFILE_SEARCH:
            url = 'https://twitter.com/i/profiles/show/' + query_string + '/timeline/with_replies'
            params['include_available_features'] = '1'
            params['include_entities'] = '1'
            params['reset_error_state'] = 'false'

        else:
            url = 'https://twitter.com/i/search/timeline'
            params['include_available_features'] = '1'
            params['include_entities'] = '1'
            params['reset_error_state'] = 'false'
            params['vertical'] = 'default'
            params['src'] = 'typd'
            params['f'] = 'tweets'
            params['lang'] = 'en'
            params['q'] = ' ' + query_string + ' since:' + data_begin + ' until:' + data_end

        params['max_position'] = refreshCursor
        resp = page.load(url, params=params, headers=h)
        try:
            r = json.loads(resp)
        except:
            print(resp.text)
            print('JSON error', url, page.pr)
            raise LoadingError

        if not r.get('inner', False):
            r['inner'] = r

        try:
            refreshCursor = r['inner']['min_position']
        except KeyError:
            print(resp.text)
            print('Key error', url, page.pr)
            raise LoadingError

        if not refreshCursor:
            break

        if not re.sub('[\n ]', '', r['inner']['items_html']):
            empty_count += 1
            if empty_count > 3:
                if data_current > data_begin:
                    print('Reduce date range')
                    date_range_change_count += 1
                    if date_range_change_count < 3:
                        data_end = data_current
                        empty_count = 0
                        continue
                break
            else:
                # print(r['inner']['new_latent_count'])
                print('Twitter server stopped. sleep 3 sec')
                time.sleep(3)
                continue
        empty_count = 0
        date_range_change_count = 0
        r = r['inner']['items_html']
        for tweet in cont(page, r, query_string):
            if tweet:
                data_current = re.sub(' \d+:\d+:\d+', '', str(tweet.date))
                yield tweet

    page.close()





# Content parser
def cont(page, r, query_string):
    r = re.sub('</span><span class="invisible">', '', r)
    try:
        tweets = PyQuery(r)('div.js-stream-tweet')
        # print(tweets.html())
    except:
        print('no div.js-stream-tweet')
        return None
    for tweetHTML in tweets:
        # pprint(tweetHTML)

        tweet = Twit()
        # tweet.c = user_tweet_count
        tweetPQ = PyQuery(tweetHTML)
        # pprint(tweetPQ)
        tweet.time_zone = ''  # time_zone
        res = re.findall('twitter-cashtag pretty-link js-nav" dir="ltr"><s>\$</s><b>(?:<strong>)?(.+?)</',
                         str(tweetPQ), re.M)
        tweet.symbols = []
        if res:
            res = list(set(res))
            for rt in res:
                if '$' + rt.upper() != query_string.upper():
                    tweet.symbols.append('$' + rt.upper())

        else:
            tweet.symbols = []

        # print(PyQuery(tweetHTML)('a.twitter-timeline-link'))
        tweet.urls = []

        flag = False
        for aa in PyQuery(tweetHTML)('a.twitter-timeline-link'):
            aaa = PyQuery(aa)

            if aaa.attr('data-expanded-url') and aaa.attr('data-expanded-url') != 'null' and aaa.attr(
                    'data-expanded-url') != 'None':
                tweet.urls.append(aaa.attr('data-expanded-url'))
                flag = True
                # print(tweetPQ("p.js-tweet-text").text())
            else:
                tweet.urls.append(aaa.attr('href'))
                # print(aaa.attr('href'),'https://'+aaa.text())
            aaa.remove()
        # print(tweetPQ("p.js-tweet-text").text())
        # if flag:
        #     raise LoadingError

        tweet.mentions_name = []
        tweet.mentions_id = []
        tweet.ment_s = []
        for aa in PyQuery(tweetHTML)('a.twitter-atreply'):
            aaa = PyQuery(aa)
            # mention={'screen_name':aaa.attr('href').replace('/',''),'id':aaa.attr('data-mentioned-user-id')}
            tweet.ment_s.append((aaa.attr('href').replace('/', ''), aaa.attr('data-mentioned-user-id')))
            tweet.mentions_name.append(aaa.attr('href').replace('/', ''))
            tweet.mentions_id.append(aaa.attr('data-mentioned-user-id'))

        usernameTweet = tweetPQ.attr("data-screen-name")

        t = tweetPQ("p.js-tweet-text").text().replace('# ', '#').replace('@ ', '@').replace('http:// ',
                                                                                            'http://').replace(
            'http://www. ', 'http://www.').replace('https://www. ', 'https://www.').replace('https:// ',
                                                                                            'https://')
        e = tweetPQ('img.Emoji')
        tweet.emoji = []
        for em in e:
            tweet.emoji.append(str(PyQuery(em).attr('aria-label').replace('Emoji: ', '')))
        # if DEBUG:
        #     print(tweet.emoji)
        txt = re.sub(r"\s+", " ", t);
        txt = re.sub('\$ (?P<s>[A-Z]{1,6}([._][A-Z]{1,2})?)', '$\g<s>', txt)

        if not re.search('<strong class="fullname">Tweet withheld</strong>', str(tweetPQ), re.M):
            try:
                retweets = int(tweetPQ("span.ProfileTweet-action--retweet span.ProfileTweet-actionCount").attr(
                    "data-tweet-stat-count").replace(",", ""));
            except AttributeError:
                print(str(tweetPQ))
                print('Attribute error in ProfileTweet-action--retweet')

                retweets = 0

            favorites = int(tweetPQ("span.ProfileTweet-action--favorite span.ProfileTweet-actionCount").attr(
                "data-tweet-stat-count").replace(",", ""));

            replyes = int(tweetPQ("span.ProfileTweet-action--reply span.ProfileTweet-actionCount").attr(
                "data-tweet-stat-count").replace(",", ""));

        dateSec = int(tweetPQ("small.time span.js-short-timestamp").attr("data-time"));

        if tweetPQ.attr('data-protected') == 'true':
            tweet.is_protected = True
        else:
            tweet.is_protected = False
        id = tweetPQ.attr("data-tweet-id")
        permalink = tweetPQ.attr("data-permalink-path")
        tweet.user_id = tweetPQ.attr('data-user-id')
        tweet.id = id
        tweet.permalink = 'https://twitter.com' + permalink
        tweet.screen_name = usernameTweet
        tweet.user_name = tweetPQ.attr('data-name')
        txt = re.sub('(?:https\://)|(?:http\://)', '', txt)
        # txt = re.sub('https\:\/\/', '', txt)
        # txt=re.sub('http\:\/\/','',txt)

        tweet.text = txt
        tweet.unixtime = dateSec
        tweet.date = datetime.fromtimestamp(dateSec)
        tweet.retweets_count = retweets
        tweet.favorites_count = favorites
        tweet.replyes_count = replyes
        # tweet.mentions = re.compile('(@\\w*)').findall(tweet.text)
        tweet.mentions = re.compile('(?:@[\w_]+)').findall(tweet.text)
        # tweet.hashtags = re.compile('(#\\w*)').findall(tweet.text)
        tweet.hashtags = re.compile('(?:\#+[\w_]+[\w\'_\-]*[\w_]+)').findall(tweet.text)

        tweet.geo = {}

        if tweetPQ.attr('data-retweeter'):
            tweet.is_retweet = True
            tweet.retweet_user_id = tweetPQ.attr('data-user-id')
            tweet.retweet_id = tweetPQ.attr('data-retweet-id')
        else:

            tweet.is_retweet = False

        tweet.lang = tweetPQ("p.js-tweet-text").attr("lang")

        tweet.is_reply = tweetPQ.attr("data-is-reply-to")
        tweet.data_conversation_id = tweetPQ.attr("data-conversation-id")
        if tweet.is_reply:
            tweet.is_reply = True
            tweet.data_conversation_id = tweetPQ.attr("data-conversation-id")
            tweet.is_reply_href = tweetPQ('a.js-user-profile-link').attr('href')
            tweet.is_reply_screen_name = tweet.is_reply_href.replace('/', '')
            tweet.is_reply_id = tweetPQ('a.js-user-profile-link').attr('data-user-id')
            if tweet.is_reply_id == tweet.user_id:  # reply to self
                tweet.is_reply_username = tweet.user_name
            else:
                tt = re.findall('<span class="username(.+?)</span>', str(tweetPQ('a.js-user-profile-link')),
                                re.S | re.M)

                try:
                    tweet.is_reply_username = re.findall('<b>(.+?)</b>', tt[0])[0]
                except IndexError:
                    print('ERROR', tweetPQ('a.js-user-profile-link'))
                    print(tt, tweet.is_reply_id, tweet.permalink)
                    raise LoadingError
                    # print(tweet.is_reply_username)

                    # print(tweet.is_reply_href,tweet.is_reply_screen_name,tweet.is_reply_id,tweet.is_reply_username)
        else:
            tweet.is_reply = False
            # tweet.data_conversation_id = ''
            tweet.is_reply_href = ''
            tweet.is_reply_screen_name = ''
            tweet.is_reply_id = ''
            tweet.is_reply_username = ''

        tweet.likes = None
        tweet.user_tweet_count = None
        tweet.user_following_count = None
        tweet.user_followers_count = None
        tweet.user_created = None
        tweet.is_verified = None
        tweet.website = None
        tweet.user_location = None
        if settings.ISUSERPROFILE:
            r = get_user_profile(usernameTweet, page, tweet)
            if r:
                tweet = r

        # tweet.user_created,tweet.user_followers_count,tweet.user_following_count,tweet.user_tweet_count=get_user_profile(tweet.screen_name,ses,pr)
        # geolocation
        tweet.location_name = None
        tweet.location_id = None
        if settings.ISLOCATION:
            url = 'https://twitter.com/' + tweet.screen_name + '/status/' + str(
                tweet.data_conversation_id) + '?conversation_id=' + str(tweet.data_conversation_id)
            j = get_s(page, url, tweet.screen_name, important=False)

            if j:
                tweet_status = PyQuery(j['page'])('a.js-geo-pivot-link')
                if tweet_status:
                    # print(tweet_status.text(), tweet_status.attr('data-place-id'))
                    tweet.location_name = tweet_status.text()
                    tweet.location_id = tweet_status.attr('data-place-id') if tweet_status.attr('data-place-id') else ''

        yield tweet

    return None



# Get user profiles function
def get_user_profile(usernameTweet, page, tweet):
    url = 'https://twitter.com/' + usernameTweet
    j = get_s(page, url, '')  # query_string)
    if j:
        if j['init_data']:
            if j['init_data']['profile_user']:
                # print(json.dumps(j['init_data']['profile_user'], indent=4))
                # exit()
                .120
                tweet.likes = j['init_data']['profile_user']['favourites_count']
                tweet.user_tweet_count = j['init_data']['profile_user']['statuses_count']
                tweet.user_listed_count = j['init_data']['profile_user']['listed_count']
                tweet.user_description = j['init_data']['profile_user']['description']
                tweet.user_timezone = j['init_data']['profile_user']['time_zone']
                tweet.utc_offset = j['init_data']['profile_user']['utc_offset']
                tweet.user_following_count = j['init_data']['profile_user']['friends_count']
                tweet.user_followers_count = j['init_data']['profile_user']['followers_count']
                tweet.user_created = j['init_data']['profile_user']['created_at']
                tweet.is_verified = j['init_data']['profile_user']['verified']
                tweet.website = j['init_data']['profile_user']['url']
                tweet.username = j['init_data']['profile_user']['name']
                tweet.utc_offset = j['init_data']['profile_user']['utc_offset']
                if j['init_data']['profile_user'].get('location', False):
                    tweet.user_location = j['init_data']['profile_user']['location']
                else:
                    tweet.user_location = ''
                return tweet
    return None


def get_s(page, url, screen_name, important=True):
    h = {'X-Requested-With': 'XMLHttpRequest', 'x-overlay-request': 'true', 'x-previous-page-name': 'profile'}
    resp = page.load(url, headers=h, important=important)
    j = resp
    if resp:
        try:
            j = json.loads(resp)
        except:
            print(url, resp)
            # print('AWAM: JSON popup', str(i), screen_name, url)
            print('Error limit exceeded ', resp.status_code, resp.text)

            if important:
                raise LoadingError
            else:
                return None

    return j


def get_symbols(s):
    s = s.upper()
    res = re.findall('(\$[A-Z]{1,6}([._][A-Z]{1,2})?)', s, re.M)
    if res:
        r = list(map(lambda x: x[0], res))
        r = list(set(r))
        return r
    else:
        return []

# Full scrape function
def scra(query, i, proxy, lock, session):
    def tokenize(s):
        return tokens_re.findall(s)

    def preprocess(s, lowercase=False):
        tokens = tokenize(s)
        if lowercase:
            tokens = [token if emoticon_re.search(token) else token.lower() for token in tokens]
        return tokens

    q = query[4]
    fieldnames = ["QueryStartDate", "QueryEndDate", "Query", "DateOfActivity", "UserScreenName", "Keyword",
                  "Location", "Website", "DateJoined", "IsMention", "UserID", "TimeOfActivity", "Hashtags",
                  "Re_tweet", "NumberOfReplies", "NumberOfRe_tweets", "NumberOfFavorites", "Tweet", "tweet_id",
                  "tweet_url", "is_verified", "Urls", "UserFollowersCount", "UserFollowingCount", "UserTweetsCount",
                  "LikesCount", "CashtagSymbols", "user_location", "permno"]

    count = 0
    # for t in chain(get_retweets(query, proxy=proxy),
    #                get_retweets(query, proxy=proxy, nativeretweets=True)):
    ttm = time.time()
    tweet_list = []
    page = Page(proxy)
    # Loop trough search results...
    for t in get_new_search(page, query):
        # pprint(t)
        # d = str(datetime.fromtimestamp(t.unixtime))
        # t1 = datetime.strptime(query[2], '%Y-%m-%d %H:%M:%S')
        # t2 = datetime.strptime(query[3], '%Y-%m-%d %H:%M:%S')

        # if int(t1.timestamp()) > int(t.unixtime):
        #     continue
        # if int(t.unixtime) > int(t2.timestamp()):
        #     continue

        data = {}
        data['permno'] = q
        data['user_location'] = t.user_location
        data['LikesCount'] = t.likes
        data['Website'] = t.website
        data['QueryStartDate'] = query[2]
        data['QueryEndDate'] = query[3]
        data['Query'] = query[1]
        data['Keyword'] = query[1]
        data['TimeOfActivity'] = time.strftime('%H:%M:%S', time.localtime(t.unixtime))
        data['DateOfActivity'] = time.strftime('%d/%m/%Y', time.localtime(t.unixtime))
        data['tweet_id'] = t.id
        data['tweet_url'] = t.permalink
        data['UserID'] = t.user_id
        data['UserScreenName'] = t.screen_name
        data['UserName'] = t.user_name
        data['TimeZone'] = t.time_zone
        data['UserTweetsCount'] = t.user_tweet_count
        data['UserFollowersCount'] = t.user_followers_count
        data['UserFollowingCount'] = t.user_following_count
        data['NumberOfFavorites'] = t.favorites_count
        data['NumberOfRe_tweets'] = t.retweets_count
        data['NumberOfReplies'] = t.replyes_count
        data['Re_tweet'] = t.is_retweet
        data['is_verified'] = t.is_verified
        # data['isProtected'] = t.is_protected
        data['isReply'] = t.is_reply
        data['ReplyTweetID'] = t.data_conversation_id
        data['ReplyUserId'] = t.is_reply_id
        # data['ReplyScreenName'] = t.is_reply_screen_name
        # data['Lang'] = t.lang
        data['Hashtags'] = ' '.join(t.hashtags)
        data['Urls'] = ', '.join(t.urls)
        data['CashtagSymbols'] = '\n'.join(t.symbols)
        # data['Mentions_id'] = '\n'.join(t.mentions_id)
        data['IsMention'] = ' '.join(t.mentions_name)
        data['Location'] = t.location_name
        # data['Location ID'] = t.location_id
        data['DateJoined'] = dateparser.parse(t.user_created) if t.user_created else None
        data['Tweet'] = t.text

        if t.utc_offset:
            if t.utc_offset / 3600 >= 0:
                data['TimeZoneUTC'] = 'UTC+' + str(int(t.utc_offset / 3600))
            else:
                data['TimeZoneUTC'] = 'UTC' + str(int(t.utc_offset / 3600))
        else:
            data['TimeZoneUTC'] = None


        tokens = preprocess(t.text)
        cashtags = [term for term in tokens if term.startswith('$') and len(term) > 1]
        hashtags = [term for term in tokens if term.startswith('#') and len(term) > 1]
        mentions = [term for term in tokens if term.startswith('@') and len(term) > 1]
        urls = [term for term in tokens if term.startswith('http') and len(term) > 4]

        tweet_list.append(data)
        # pprint(data)

        print(query, count, t.date)
        # print(data)
        if settings.ISUSERPROFILE:
            pass
            # dd = re.findall('\w+ (\w+) (\d+) \d+:\d+:\d+ \+\d+ (\d+)', data['DateJoined'])
            # try:
            #     date_joined = datetime.strptime(' '.join(dd[0]), '%b %d %Y')
            # except IndexError:
            #     print(data['DateJoined'], dd)
            #     exit()
        else:
            date_joined = None
            t.user_listed_count = None
            t.username = t.user_name
            t.user_timezone = None
            t.user_description = None
            data['Website'] = None
            data['is_verified'] = None

        if session.query(Tweet).filter_by(tweet_id=data['tweet_id']).first():
            continue

        user = session.query(User).filter_by(user_id=data['UserID']).first()
        if not user:
            user_count = UserCount(follower=data['UserFollowersCount'],
                                   following=data['UserFollowingCount'],
                                   tweets=data['UserTweetsCount'],
                                   likes=data['NumberOfFavorites'],
                                   lists=t.user_listed_count)
            session.add(user_count)
            user = User(user_id=data['UserID'],
                        twitter_handle=data['UserScreenName'][:120],
                        user_name=data['UserName'][:120],
                        location=data['user_location'][:255] if data['user_location'] else None,
                        date_joined=data['DateJoined'],
                        timezone=data['TimeZoneUTC'],
                        website=data['Website'][:255] if data['Website'] else None,
                        user_intro=t.user_description[:255] if t.user_description else None,
                        verified=data['is_verified'])
            user.counts.append(user_count)
            session.add(user)
            try:
                session.commit()
            except sqlalchemy.exc.IntegrityError as err:
                if re.match("(.*)Duplicate entry(.*)for key 'PRIMARY'(.*)", err.args[0]):
                    print('ROLLBACK USER')
                    session.rollback()
            except Exception as e:
                print(e)
                raise

        twit = session.query(Tweet).filter_by(tweet_id=data['tweet_id']).first()
        if not twit:
            twit = Tweet(tweet_id=data['tweet_id'],
                         date=datetime.strptime(data['DateOfActivity'], '%d/%m/%Y'),
                         time=data['TimeOfActivity'],
                         timezone=data['TimeZone'][:10] if t.user_timezone else None,
                         retweet_status=data['Re_tweet'],
                         text=data['Tweet'],
                         location=data['Location'][:255] if data['Location'] else None,
                         permalink=data['tweet_url'] if data['tweet_url'] else None,
                         emoticon=','.join(t.emoji) if t.emoji else None)
            tweet_count = TweetCount(reply=data['NumberOfReplies'],
                                     favorite=data['NumberOfFavorites'],
                                     retweet=t.retweets_count)

        if t.is_reply and settings.ISREPLY:
            url = 'https://twitter.com/' + t.screen_name + '/status/' + str(
                data['tweet_id']) + '?conversation_id=' + str(t.data_conversation_id)
            r1 = page.load(url)
            for tw in PyQuery(r1)('li.js-stream-item'):
                r_date_time = PyQuery(tw)('span.js-short-timestamp').attr('data-time')
                r_text = PyQuery(tw)('p.js-tweet-text').text()
                data_user_id = PyQuery(tw)('div.js-stream-tweet').attr('data-user-id')
                reply_item_id = PyQuery(tw).attr('data-item-id')
                e = PyQuery(tw)('img.Emoji')
                r_emoji = []
                for em in e:
                    r_emoji.append(str(PyQuery(em).attr('aria-label').replace('Emoji: ', '')))
                if re.search('show_more_button', r1,
                             re.M):  # PyQuery(r1)('li.ThreadedConversation-moreReplies').attr('data-expansion-url')
                    u = 'https://twitter.com/i/' + t.screen_name + '/conversation/' + str(
                        t.data_conversation_id) + '?include_available_features=1&include_entities=1&max_position=DAACDwABCgAAAA0NGVYFElfQAQ0ZUm161tABDRlmNLsXUAENGVN_UddwAA0ZYae2FgAADRlT_t6XcAMNGVX62haQBg0ZUn9bF-ABDRlYig7W4AANGVM58xfgAA0ZUxcJ15ABDRlUJzMW0AENGW38gpbQAAgAAwAAAAECAAQAAAA&reset_error_state=false'
                print('This is an error on purpuse. The reply_id is', data['tweet_id'])
                if not session.query(Reply).filter_by(reply_id=data['tweet_id']).first():
                    parent_twit = session.query(Tweet).filter_by(tweet_id=t.data_conversation_id).first()
                    TimeOfActivity = time.strftime('%H:%M:%S', time.localtime(int(r_date_time)))
                    DateOfActivity = time.strftime('%d/%m/%Y', time.localtime(int(r_date_time)))
                    reply = Reply(reply_id=reply_item_id,
                                  reply_user_id=data_user_id,
                                  date=datetime.strptime(DateOfActivity, '%d/%m/%Y'),
                                  time=TimeOfActivity,
                                  timezone=None,
                                  text=r_text,
                                  emoticon=','.join(r_emoji) if r_emoji else None)
                    session.add(reply)
                    if parent_twit:
                        parent_twit.replies.append(reply)
                    else:
                        twit.replies.append(reply)

                try:
                    session.commit()
                except sqlalchemy.exc.IntegrityError as err:
                    if re.match("(.*)Duplicate entry(.*)for key 'PRIMARY'(.*)", err.args[0]):
                        print('ROLLBACK REPLY')
                        session.rollback()
                except Exception as e:
                    print(e)
                    raise

        if not session.query(TweetHashtags).filter_by(tweet_id=data['tweet_id']).first():
            for hash_s in hashtags:
                tweet_hashtag = TweetHashtags(hashtags=hash_s[:45])
                twit.hash_s.append(tweet_hashtag)
                session.add(tweet_hashtag)

        if not session.query(TweetUrl).filter_by(tweet_id=data['tweet_id']).first():
            for url_s in urls:
                tweet_url = TweetUrl(url=url_s[:255])
                twit.url_s.append(tweet_url)
                session.add(tweet_url)

        if not session.query(TweetCashtags).filter_by(tweet_id=data['tweet_id']).first():
            for cash_s in cashtags:
                tweet_cashtag = TweetCashtags(cashtags=cash_s[:45])
                twit.cash_s.append(tweet_cashtag)
                session.add(tweet_cashtag)

        if not session.query(TweetMentions).filter_by(tweet_id=data['tweet_id']).first():
            for ment_s in t.ment_s:
                tweet_mentions = TweetMentions(mentions=ment_s[0][:45], user_id=ment_s[1])
                twit.ment_s.append(tweet_mentions)
                session.add(tweet_mentions)
        user.tweets.append(twit)
        twit.counts.append(tweet_count)
        session.add(tweet_count)
        session.add(twit)
        try:
            session.commit()
        except sqlalchemy.exc.IntegrityError as err:
            if re.match("(.*)Duplicate entry(.*)for key 'PRIMARY'(.*)", err.args[0]):
                print('ROLLBACK USER')
                session.rollback()
        except Exception as e:
            print(e)
            raise
        count += 1

    lock.acquire()
    if count > 0:

        with open('report.csv', 'a') as f:  # Add reports to this file
            data = {}
            fdnames = ['time', 'query_name', 'number', ]
            writer = csv.DictWriter(f, lineterminator='\n', fieldnames=fdnames, dialect='excel', quotechar='"',
                                    quoting=csv.QUOTE_ALL)

            data['time'] = time.strftime('%Y-%m-%d %H:%M:%S')
            data['query_name'] = query[1]
            data['number'] = count

            writer.writerow(data)
        lock.release()
        return count
    else:
        with open('error.csv', 'a') as f:  # Add errors to this file
            data = {}
            fdnames = ['time', 'query_name', 'number', ]
            writer = csv.DictWriter(f, lineterminator='\n', fieldnames=fdnames, dialect='excel', quotechar='"',
                                    quoting=csv.QUOTE_ALL)

            data['time'] = time.strftime('%Y-%m-%d %H:%M:%S')
            data['query_name'] = query[1]
            data['number'] = count

            writer.writerow(data)
        lock.release()
        return False


def compare_90(string1, string2):
    if len(string1) == 0 and len(string2) == 0:
        return True
    if len(string1) >= len(string2):
        s1 = string1
        s2 = string2
    else:
        s1 = string2
        s2 = string1

    if len(s2) == 0 and len(s2) != len(s1):
        return False

    if float(len(s1) / len(s2)) > float(100 / COEF):
        return False

    mismatch_count = 0.0
    treshold = float(len(s1) * (1 - (COEF / 100)))
    for i in range(len(s2)):

        if s1[i] != s2[i]:
            mismatch_count += 1.0
            if mismatch_count > treshold:
                return False

    return True


def scrape_query(user_queue, proxy, lock, pg_dsn):
    db_engine = create_engine(pg_dsn, pool_size=1)
    add_engine_pidguard(db_engine)
    DstSession = sessionmaker(bind=db_engine, autoflush=False)
    session = DstSession()

    active = True
    while not user_queue.empty():
        query, i = user_queue.get()

        print('START', i, proxy, query)
        # TODO filter:nativeretweets
        try:
            res = scra(query, i, proxy, lock, session=session)
        except LoadingError:
            print('LoadingError except')
            return False
        if not res:
            print('     SCRAP_USER Error in', query, i)
            with open('error_list.txt', 'a') as f:
                f.write(query[0] + '\n')
        else:
            print(datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ' ENDED', i, proxy, query, res)
    return True


def get_up(fname, proxy):
    n = 0
    fieldnames = ["QueryStartDate", "QueryEndDate", "Query", "DateOfActivity", "UserScreenName", "Keyword",
                  "Location", "Website", "DateJoined", "IsMention", "UserID", "TimeOfActivity", "Hashtags",
                  "Re_tweet", "NumberOfReplies", "NumberOfRe_tweets", "NumberOfFavorites", "Tweet", "tweet_id",
                  "tweet_url", "is_verified", "Urls", "UserFollowersCount", "UserFollowingCount", "UserTweetsCount",
                  "LikesCount", "CashtagSymbols", "user_location", "permno"]
    # print('Processing file {} ...'.format(fname))
    with open(fname, 'r', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        recordlist = []
        count = 0
        for row in reader:
            recordlist.append(row)
            count += 1
    # print('   Read {} rows'.format(count))

    page = Page(proxy)
    new_list = []
    i = 0
    user_profiles = {}
    tt = time.time()
    for row in recordlist:
        tweet = Tweet()
        try:
            if user_profiles.get(row['UserScreenName'], None):
                r = Tweet()
                rt = user_profiles[row['UserScreenName']]
                r.user_location = rt['user_location']
                r.user_tweet_count = rt['UserTweetsCount']
                r.user_following_count = rt['UserFollowingCount']
                r.user_followers_count = rt['UserFollowersCount']
                r.likes = rt['LikesCount']
                r.user_created = rt['DateJoined']
                r.website = rt['Website']
                r.is_verified = rt['is_verified']

            else:
                try:
                    r = get_user_profile(row['UserScreenName'], page, tweet)
                except LoadingError:
                    print('Skip user', row['UserScreenName'])
                    r = None

            if r:
                tweet = r

                # print(tweet.likes,tweet.user_tweet_count,tweet.user_following_count,tweet.user_followers_count,
                # tweet.user_created,tweet.is_verified,tweet.website,tweet.user_location)
                rt = {}
                rt['user_location'] = row['user_location'] = tweet.user_location
                rt['UserTweetsCount'] = row['UserTweetsCount'] = tweet.user_tweet_count
                rt['UserFollowingCount'] = row['UserFollowingCount'] = tweet.user_following_count
                rt['UserFollowersCount'] = row['UserFollowersCount'] = tweet.user_followers_count
                rt['LikesCount'] = row['LikesCount'] = tweet.likes
                rt['DateJoined'] = row['DateJoined'] = tweet.user_created
                rt['Website'] = row['Website'] = tweet.website
                rt['is_verified'] = row['is_verified'] = tweet.is_verified
                user_profiles[row['UserScreenName']] = rt
            new_list.append(row)
            i += 1

            k = 50
            if i % k == 0:
                print('  {} Loaded {} user profiles '.format(fname, i))
                tt = time.time()
        except KeyError as e:
            print('File {} has the wrong content,skipped'.format(fname))
            break
    else:
        with open(fname, 'w', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, extrasaction='ignore', restval='', lineterminator='\n',
                                    fieldnames=fieldnames,
                                    dialect='excel', quotechar='"',
                                    quoting=csv.QUOTE_ALL)
            writer.writeheader()
            writer.writerows(new_list)
        # print('Write {} rows'.format(len(new_list)))
        n = len(new_list)

    return n


def get_pup(user_queue, proxy):
    while not user_queue.empty():
        fname, i = user_queue.get()

        print('START', i, proxy, fname)

        res = get_up(fname, proxy)

        print('ENDED', i, proxy, fname, res)
    return True


def add_engine_pidguard(engine):
     """Add multiprocessing guards.

    Forces a connection to be reconnected if it is detected
    as having been shared to a sub-process.

    """
    @event.listens_for(engine, "connect")
    def connect(dbapi_connection, connection_record):
        connection_record.info['pid'] = os.getpid()  # Returns the current process id

    @event.listens_for(engine, "checkout")
    def checkout(dbapi_connection, connection_record, connection_proxy):
        pid = os.getpid()
        if connection_record.info['pid'] != pid:
            # substitute log.debug() or similar here as desired
            warnings.warn(
                "Parent process %(orig)s forked (%(newproc)s) with an open "
                "database connection, "
                "which is being discarded and recreated." %
                {"newproc": pid, "orig": connection_record.info['pid']})
            connection_record.connection = connection_proxy.connection = None
            raise exc.DisconnectionError(
                "Connection record belongs to pid %s, "
                "attempting to check out in pid %s" %
                (connection_record.info['pid'], pid)
            )


# start of main program
if __name__ == '__main__':
    # making dict with postgres login data
    pg_config = {'username': settings.PG_USER, 'password': settings.PG_PASSWORD, 'database': settings.PG_DBNAME,
                 'host': settings.DB_HOST}
    # postgres connection string
    pg_dsn = "postgresql+psycopg2://{username}:{password}@{host}:5432/{database}".format(**pg_config)
    # Construct a base class for declarative class definitions.
    Base = declarative_base()
    db_engine = create_engine(pg_dsn)
    add_engine_pidguard(db_engine)
    # container object that keeps together many different features of a database (or multiple databases) being described.
    pg_meta = MetaData(bind=db_engine, schema="fintweet")


    # Reflect destination tables
    print('I am here !')
    # class User(Base):
    #     __table__ = Table('user', pg_meta, autoload=True)
    #     tweets = relationship('Tweet')
    #     counts = relationship('UserCount')
    #
    #
    # class UserCount(Base):
    #     __table__ = Table('user_count', pg_meta, autoload=True)
    #
    #
    # class TweetCount(Base):
    #     __table__ = Table('tweet_count', pg_meta, autoload=True)
    #
    #
    # class TweetMentions(Base):
    #     __table__ = Table('tweet_mentions', pg_meta, autoload=True)
    #
    #
    # class TweetCashtags(Base):
    #     __table__ = Table('tweet_cashtags', pg_meta, autoload=True)
    #
    #
    # class TweetHashtags(Base):
    #     __table__ = Table('tweet_hashtags', pg_meta, autoload=True)
    #
    #
    # class TweetUrl(Base):
    #     __table__ = Table('tweet_url', pg_meta, autoload=True)
    #
    #
    # class Tweet(Base):
    #     __table__ = Table('tweet', pg_meta, autoload=True)
    #     counts = relationship('TweetCount')
    #     ment_s = relationship('TweetMentions')
    #     cash_s = relationship('TweetCashtags')
    #     hash_s = relationship('TweetHashtags')
    #     url_s = relationship('TweetUrl')


    # class Reply(Base):
    #     __tablename__ = 'reply'
    #     reply_id = Column(BIGINT, primary_key=True)
    #     tweet_id = Column(BIGINT, ForeignKey('tweet.tweet_id'))
    #     reply_user_id = Column(BIGINT)
    #     date = Column(DATE)
    #     time = Column(TIME(6))
    #     timezone = Column(VARCHAR(10))
    #     text = Column(TEXT)
    #     emoticon = Column(TEXT)

    # config = load_config()

    DstSession = sessionmaker(bind=db_engine, autoflush=False)
    dstssn = DstSession()

    if True:  # settings.TWEETS:
        try:
            command = sys.argv[1]
            print(command)
        except IndexError:
            command = ''

        if command == 'location':
            ISLOCATION = True
        else:
            ISLOCATION = False

        user_queue = ThreadQueue()

        # load excel file for input
        fname = 'word_list.xlsx'
        wb = load_workbook(fname)
        ws = wb.active
        ii = i = 2

        while True:
            if not ws.cell(row=i, column=1).value:
                break
            t1 = str(ws.cell(row=i, column=4).value).lower().strip(' ')
            t2 = str(ws.cell(row=i, column=5).value).lower().strip(' ')
            t1 = re.sub(' 00:00:00', '', t1)
            t2 = re.sub(' 00:00:00', '', t2)
            permno = str(ws.cell(row=i, column=1).value).lower().strip(' ')
            query = str(ws.cell(row=i, column=2).value).lower().strip(' '), \
                    str(ws.cell(row=i, column=3).value).lower().strip(' '), \
                    t1, t2, permno
            print(query)

            user_queue.put((query, i))
            i += 1

        pool = ThreadPool(len(settings.PROXY_LIST))
        lock = Lock()
        pool.map(lambda x: (scrape_query(user_queue, x, lock, pg_dsn)), settings.PROXY_LIST)
    else:

        path = '.'
        csv_files = [f for f in os.listdir(path) if f.endswith('.csv')]
        print('Collecting user data from {} files'.format(len(csv_files)))

        # pool = ThreadPool(len(settings.PROXY_LIST))
        # user_queue = ThreadQueue()

        user_queue = Queue()
        i = 0
        for fname in csv_files:
            i += 1
            user_queue.put((fname, i))
        # pool.map(lambda x: (get_pup(user_queue, x)), settings.PROXY_LIST)

        pp = []
        for s in settings.PROXY_LIST:
            p = Process(target=get_pup, args=(user_queue, s))
            p.start()
            pp.append(p)
        #
        for p in pp:
            p.join()
